#!/usr/bin/env python3
# SPDX-License-Identifier: MIT
"""Unified bank extract — auto-detect adapter, enforce balance proof where possible.

Adapters:
  - maybank_islamic_pdf  (existing script)
  - generic_csv          (flexible headers)
  - cimb_csv             (CIMB Clicks / business CSV exports)

Usage:
  python3 scripts/extract_bank.py --input ./statements --output ./bank.xlsx
  python3 scripts/extract_bank.py --input ./bank.csv --also-json ./txns.json --client-slug acme
  python3 scripts/extract_bank.py --detect-only ./statements
"""
from __future__ import annotations

import argparse
import csv
import json
import re
import subprocess
import sys
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
MAYBANK_SCRIPT = ROOT / "scripts" / "extract_maybank_islamic_pdf.py"

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    Workbook = None  # type: ignore


def money(x) -> Decimal:
    return Decimal(str(x)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def collect_files(path: Path) -> list[Path]:
    if path.is_file():
        return [path]
    files: list[Path] = []
    for p in sorted(path.rglob("*")):
        if p.suffix.lower() in (".pdf", ".csv", ".tsv", ".txt") and p.is_file():
            files.append(p)
    return files


def sniff_pdf_bank(text: str) -> str:
    t = text[:8000]
    if re.search(r"Maybank\s+Islamic|Maybank2u|MAYBANK ISLAMIC", t, re.I):
        return "maybank_islamic_pdf"
    if re.search(r"\bCIMB\b|CIMB Bank|CIMB Clicks", t, re.I):
        return "cimb_pdf_unsupported"
    if re.search(r"Public Bank|PBe\b", t, re.I):
        return "public_bank_pdf_unsupported"
    if re.search(r"\bRHB\b", t, re.I):
        return "rhb_pdf_unsupported"
    if re.search(r"Hong Leong", t, re.I):
        return "hong_leong_pdf_unsupported"
    if re.search(r"HSBC", t, re.I):
        return "hsbc_pdf_unsupported"
    return "unknown_pdf"


def sniff_csv_bank(header: list[str], sample_rows: list[dict]) -> str:
    h = " ".join(header).lower()
    if "cimb" in h or any("cimb" in str(v).lower() for r in sample_rows for v in r.values()):
        return "cimb_csv"
    # CIMB-style debit/credit columns without brand
    if ("debit" in h and "credit" in h) or ("withdrawal" in h and "deposit" in h):
        return "cimb_csv" if "transaction description" in h or "txn" in h else "generic_csv"
    return "generic_csv"


def detect(path: Path) -> list[dict]:
    reports = []
    for f in collect_files(path):
        if f.suffix.lower() == ".pdf":
            try:
                import pdfplumber
            except ImportError:
                reports.append({"file": str(f), "adapter": "need_pdfplumber", "ok": False})
                continue
            with pdfplumber.open(f) as pdf:
                text = "\n".join((p.extract_text() or "") for p in pdf.pages[:2])
            adapter = sniff_pdf_bank(text)
            reports.append(
                {
                    "file": str(f),
                    "adapter": adapter,
                    "ok": adapter == "maybank_islamic_pdf",
                    "hint": None
                    if adapter == "maybank_islamic_pdf"
                    else "Export CSV from internet banking, or request a PDF adapter.",
                }
            )
        else:
            with f.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
                reader = csv.DictReader(fh)
                rows = []
                for i, row in enumerate(reader):
                    rows.append(row)
                    if i >= 4:
                        break
                header = list(reader.fieldnames or [])
            adapter = sniff_csv_bank(header, rows)
            reports.append({"file": str(f), "adapter": adapter, "ok": True, "header": header})
    return reports


def pick(row: dict, *names: str):
    lower = {re.sub(r"\s+", " ", k.lower().strip()): v for k, v in row.items() if k}
    for n in names:
        if n in lower and str(lower[n]).strip() != "":
            return str(lower[n]).strip()
    return None


def parse_amount(s: str) -> Decimal:
    s = s.replace(",", "").replace("RM", "").replace("myr", "").strip()
    s = re.sub(r"[^\d.\-]", "", s)
    if not s or s in (".", "-"):
        return Decimal("0.00")
    return money(s)


def parse_date(s: str) -> str:
    s = s.strip()
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2})$", s)
    if m:
        d, mo, y = m.groups()
        yi = int(y)
        yi += 2000 if yi < 70 else 1900
        return f"{yi}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return s[:10]
    raise ValueError(f"unrecognized date: {s}")


def parse_csv_file(path: Path, adapter: str) -> tuple[list[dict], dict]:
    rows_out: list[dict] = []
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        reader = csv.DictReader(fh)
        for i, row in enumerate(reader, start=2):
            desc = pick(
                row,
                "description",
                "transaction description",
                "txn description",
                "particulars",
                "details",
                "narrative",
            )
            date_raw = pick(
                row,
                "date",
                "transaction date",
                "txn date",
                "value date",
                "posting date",
                "entry date",
            )
            if not date_raw or not desc:
                continue
            # skip opening/closing markers
            dl = desc.lower()
            if "opening balance" in dl or "closing balance" in dl or "begin balance" in dl:
                continue

            debit = pick(row, "debit", "withdrawal", "money out", "dr")
            credit = pick(row, "credit", "deposit", "money in", "cr")
            amount_raw = pick(row, "amount", "transaction amount")
            direction = pick(row, "direction", "type", "dr/cr")

            amt = Decimal("0.00")
            dirn = None
            if debit and parse_amount(debit) > 0:
                amt = parse_amount(debit)
                dirn = "outflow"
            elif credit and parse_amount(credit) > 0:
                amt = parse_amount(credit)
                dirn = "inflow"
            elif amount_raw:
                amt = abs(parse_amount(amount_raw))
                if direction:
                    dl2 = direction.lower()
                    if dl2 in ("inflow", "credit", "cr", "deposit", "c"):
                        dirn = "inflow"
                    elif dl2 in ("outflow", "debit", "dr", "withdrawal", "d", "payment"):
                        dirn = "outflow"
                if dirn is None:
                    dirn = "outflow" if amount_raw.strip().startswith("-") else None

            if amt <= 0 or not dirn:
                continue

            bal_raw = pick(row, "balance", "running balance", "ledger balance")
            bal = parse_amount(bal_raw) if bal_raw else None

            rows_out.append(
                {
                    "date": parse_date(date_raw),
                    "description": desc,
                    "amount": float(amt),
                    "direction": dirn,
                    "running_balance": float(bal) if bal is not None else None,
                    "source_file": str(path.name),
                    "source_ref": f"row:{i}",
                    "adapter": adapter,
                }
            )

    meta = {
        "file": path.name,
        "adapter": adapter,
        "txn_count": len(rows_out),
        "line_balance_ok": None,
        "errors": [],
    }
    # Optional running-balance proof when balances present
    with_bal = [r for r in rows_out if r.get("running_balance") is not None]
    if len(with_bal) >= 2:
        ok = True
        for a, b in zip(with_bal, with_bal[1:]):
            signed = money(b["amount"]) if b["direction"] == "inflow" else -money(b["amount"])
            expected = money(a["running_balance"]) + signed
            if abs(expected - money(b["running_balance"])) > money("0.02"):
                ok = False
                meta["errors"].append(
                    f"balance break after {b['date']} {b['description'][:40]}: "
                    f"expected {expected} got {b['running_balance']}"
                )
                break
        meta["line_balance_ok"] = ok
    return rows_out, meta


def write_excel(rows: list[dict], metas: list[dict], path: Path) -> None:
    if Workbook is None:
        raise SystemExit("ERROR: pip install openpyxl")
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"
    headers = [
        "date",
        "description",
        "amount",
        "direction",
        "running_balance",
        "source_file",
        "source_ref",
        "adapter",
    ]
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    ws2 = wb.create_sheet("Meta")
    ws2.append(["file", "adapter", "txn_count", "line_balance_ok", "errors"])
    for m in metas:
        ws2.append(
            [
                m.get("file"),
                m.get("adapter"),
                m.get("txn_count"),
                m.get("line_balance_ok"),
                "; ".join(m.get("errors") or []),
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def rows_to_schema(rows: list[dict], metas: list[dict], client_slug: str) -> dict:
    txns = []
    for i, r in enumerate(rows, start=1):
        txns.append(
            {
                "id": f"txn-{i:04d}",
                "date": r["date"],
                "description": r["description"],
                "amount": r["amount"],
                "direction": r["direction"],
                "bank_account_id": "bank-001",
                "running_balance": r.get("running_balance"),
                "source_file": r.get("source_file"),
                "source_ref": r.get("source_ref"),
                "extraction_adapter": r.get("adapter"),
            }
        )
    return {
        "schema_version": "0.0.1",
        "client_slug": client_slug,
        "currency": "MYR",
        "bank_accounts": [
            {
                "id": "bank-001",
                "name": "Primary bank",
                "account_number_last4": None,
            }
        ],
        "transactions": txns,
        "extraction_meta": metas,
    }


def run_maybank(files: list[Path], output: Path, also_json: Path | None, client_slug: str, fail: bool) -> int:
    # Feed folder or single file to existing extractor
    # If mixed, only PDFs
    pdfs = [f for f in files if f.suffix.lower() == ".pdf"]
    if not pdfs:
        return 1
    # Use parent if all same dir else temp-like: pass first parent
    inp = pdfs[0] if len(pdfs) == 1 else pdfs[0].parent
    cmd = [
        sys.executable,
        str(MAYBANK_SCRIPT),
        "--input",
        str(inp),
        "--output",
        str(output),
        "--client-slug",
        client_slug,
    ]
    if also_json:
        cmd.extend(["--also-json", str(also_json)])
    if fail:
        cmd.append("--fail-on-error")
    return subprocess.call(cmd)


def main() -> int:
    ap = argparse.ArgumentParser(description="Unified bank extract with adapter detection")
    ap.add_argument("--input", type=Path, required=True)
    ap.add_argument("--output", type=Path, default=None)
    ap.add_argument("--also-json", type=Path, default=None)
    ap.add_argument("--client-slug", default="client")
    ap.add_argument("--fail-on-error", action="store_true")
    ap.add_argument("--detect-only", action="store_true")
    args = ap.parse_args()

    if not args.input.exists():
        print(f"ERROR: not found: {args.input}", file=sys.stderr)
        return 1

    reports = detect(args.input)
    if args.detect_only:
        print(json.dumps(reports, indent=2))
        return 0 if reports else 1

    if not reports:
        print(f"ERROR: no bank files under {args.input}", file=sys.stderr)
        return 1

    files = collect_files(args.input)
    pdfs = [f for f in files if f.suffix.lower() == ".pdf"]
    csvs = [f for f in files if f.suffix.lower() in (".csv", ".tsv", ".txt")]

    # PDF path — Maybank only today
    maybank_pdfs = []
    blocked = []
    for r in reports:
        if r["adapter"] == "maybank_islamic_pdf":
            maybank_pdfs.append(Path(r["file"]))
        elif str(r["adapter"]).endswith("_pdf_unsupported") or r["adapter"] == "unknown_pdf":
            blocked.append(r)

    out_xlsx = args.output or Path(f"bank_extract_{args.client_slug}.xlsx")
    all_rows: list[dict] = []
    metas: list[dict] = []

    if maybank_pdfs and not csvs:
        code = run_maybank(maybank_pdfs, out_xlsx, args.also_json, args.client_slug, args.fail_on_error)
        return code

    if maybank_pdfs and csvs:
        # Run maybank into temp-less: run maybank for pdfs, merge csv after
        code = run_maybank(maybank_pdfs, out_xlsx, None, args.client_slug, args.fail_on_error)
        if code != 0:
            return code
        print("NOTE: PDF extract written; CSV rows will be merged into JSON if --also-json set")

    for f in csvs:
        with f.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
            reader = csv.DictReader(fh)
            header = list(reader.fieldnames or [])
            sample = []
            # re-open for full parse
        adapter = sniff_csv_bank(header, [])
        # Re-detect with content
        adapter = "cimb_csv" if adapter == "cimb_csv" else "generic_csv"
        # Brand sniff from filename
        if "cimb" in f.name.lower():
            adapter = "cimb_csv"
        rows, meta = parse_csv_file(f, adapter)
        all_rows.extend(rows)
        metas.append(meta)
        status = "PASS" if meta.get("line_balance_ok") in (True, None) else "CHECK"
        print(f"{status} {f.name} adapter={adapter} txns={meta['txn_count']} errors={len(meta.get('errors') or [])}")

    if blocked and not all_rows and not maybank_pdfs:
        for b in blocked:
            print(
                f"ERROR: no PDF adapter for {b['file']} ({b['adapter']}). {b.get('hint') or ''}",
                file=sys.stderr,
            )
        return 1

    if all_rows:
        all_rows.sort(key=lambda r: (r["date"], r.get("source_ref") or ""))
        write_excel(all_rows, metas, out_xlsx)
        print(f"Wrote {out_xlsx} ({len(all_rows)} transactions)")
        if args.also_json:
            payload = rows_to_schema(all_rows, metas, args.client_slug)
            args.also_json.parent.mkdir(parents=True, exist_ok=True)
            args.also_json.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
            print(f"Wrote {args.also_json}")
        failed = any(m.get("line_balance_ok") is False for m in metas)
        if args.fail_on_error and failed:
            return 1
        return 0

    if maybank_pdfs:
        return 0  # already written

    print("ERROR: nothing extracted", file=sys.stderr)
    return 1


if __name__ == "__main__":
    sys.exit(main())
