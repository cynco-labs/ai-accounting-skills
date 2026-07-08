#!/usr/bin/env python3
"""
Batch Excel Workbook Generator for Malaysian Accounting Engagements
Generates complete working papers from classified transaction data in a single pass.

Usage:
    python generate_workbook.py <input_json> <output_xlsx> [--entity-type TYPE]

Input JSON structure:
{
    "company_info": { "name": "...", "registration": "...", "fy_start": "...", "fy_end": "..." },
    "chart_of_accounts": { "code": {"name": "...", "type": "...", "normal": "..."} },
    "opening_balances": { "code": amount },
    "bank_transactions": [
        {"date": "YYYY-MM-DD", "description": "...", "amount": 0.00, "type": "DR|CR",
         "account_code": "...", "bank_account": "...", "je_number": "..."}
    ],
    "payroll": [
        {"month": "YYYY-MM", "employee": "...", "gross": 0.00, "epf_er": 0.00,
         "epf_ee": 0.00, "socso_er": 0.00, "socso_ee": 0.00, "eis_er": 0.00,
         "eis_ee": 0.00, "pcb": 0.00, "net_pay": 0.00}
    ],
    "fixed_assets": [
        {"description": "...", "category": "...", "date_acquired": "YYYY-MM-DD",
         "cost": 0.00, "rate": 0.00, "accumulated_depr": 0.00}
    ],
    "journal_entries": [
        {"je_number": "...", "date": "YYYY-MM-DD", "description": "...",
         "lines": [{"account_code": "...", "debit": 0.00, "credit": 0.00}]}
    ],
    "tax_computation": {
        "net_profit": 0.00, "add_backs": [{"description": "...", "amount": 0.00}],
        "less_items": [{"description": "...", "amount": 0.00}],
        "capital_allowances": {"bf": 0.00, "ia": 0.00, "aa": 0.00, "absorbed": 0.00, "cf": 0.00}
    },
    "queries": [
        {"date": "YYYY-MM-DD", "amount": 0.00, "description": "...", "question": "...", "status": "open|resolved"}
    ]
}
"""

import json
import sys
import argparse
from datetime import datetime, date
from collections import defaultdict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


HEADER_FONT = Font(name='Arial', size=11, bold=True)
NORMAL_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', size=14, bold=True)
SUBTITLE_FONT = Font(name='Arial', size=11, italic=True)
NUMBER_FORMAT = '#,##0.00'
HEADER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
THIN_BORDER = Border(
    bottom=Side(style='thin')
)
DOUBLE_BORDER = Border(
    top=Side(style='thin'),
    bottom=Side(style='double')
)


def set_column_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_header_row(ws, row, headers, widths=None):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
    if widths:
        set_column_widths(ws, widths)
    return row + 1


def write_company_info(wb, data):
    ws = wb.create_sheet("Company Info")
    info = data.get('company_info', {})

    ws.cell(row=1, column=1, value="COMPANY INFORMATION").font = TITLE_FONT
    ws.cell(row=3, column=1, value="Entity Name:").font = HEADER_FONT
    ws.cell(row=3, column=2, value=info.get('name', '')).font = NORMAL_FONT
    ws.cell(row=4, column=1, value="Registration No:").font = HEADER_FONT
    ws.cell(row=4, column=2, value=info.get('registration', '')).font = NORMAL_FONT
    ws.cell(row=5, column=1, value="Entity Type:").font = HEADER_FONT
    ws.cell(row=5, column=2, value=info.get('entity_type', '')).font = NORMAL_FONT
    ws.cell(row=6, column=1, value="Framework:").font = HEADER_FONT
    ws.cell(row=6, column=2, value=info.get('framework', '')).font = NORMAL_FONT
    ws.cell(row=7, column=1, value="Financial Year:").font = HEADER_FONT
    ws.cell(row=7, column=2, value=f"{info.get('fy_start', '')} to {info.get('fy_end', '')}").font = NORMAL_FONT
    ws.cell(row=8, column=1, value="Tax Filing:").font = HEADER_FONT
    ws.cell(row=8, column=2, value=info.get('tax_form', '')).font = NORMAL_FONT
    ws.cell(row=10, column=1, value="Prepared by:").font = HEADER_FONT
    ws.cell(row=10, column=2, value=info.get('prepared_by', 'Hazli Johar & Co. (NF1932)')).font = NORMAL_FONT
    ws.cell(row=11, column=1, value="Date Prepared:").font = HEADER_FONT
    ws.cell(row=11, column=2, value=datetime.now().strftime('%d/%m/%Y')).font = NORMAL_FONT

    set_column_widths(ws, [20, 40])


def write_chart_of_accounts(wb, data):
    ws = wb.create_sheet("Chart of Accounts")
    coa = data.get('chart_of_accounts', {})

    ws.cell(row=1, column=1, value="CHART OF ACCOUNTS").font = TITLE_FONT
    row = write_header_row(ws, 3, ["Code", "Account Name", "Type", "Normal Balance"], [10, 40, 15, 15])

    for code in sorted(coa.keys()):
        acc = coa[code]
        ws.cell(row=row, column=1, value=code).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=acc.get('name', '')).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=acc.get('type', '')).font = NORMAL_FONT
        ws.cell(row=row, column=4, value=acc.get('normal', '')).font = NORMAL_FONT
        row += 1


def write_bank_transactions(wb, data):
    ws = wb.create_sheet("Bank Transactions")
    transactions = data.get('bank_transactions', [])

    ws.cell(row=1, column=1, value="BANK TRANSACTIONS").font = TITLE_FONT
    headers = ["Date", "Description", "Debit (In)", "Credit (Out)", "Balance", "Account Code", "Account Name"]
    row = write_header_row(ws, 3, headers, [12, 40, 14, 14, 14, 12, 30])

    coa = data.get('chart_of_accounts', {})
    running_balance = data.get('opening_balances', {}).get('1000', 0)

    ws.cell(row=row, column=1, value="").font = NORMAL_FONT
    ws.cell(row=row, column=2, value="Opening Balance").font = Font(name='Arial', size=10, italic=True)
    ws.cell(row=row, column=5, value=running_balance).font = NORMAL_FONT
    ws.cell(row=row, column=5).number_format = NUMBER_FORMAT
    row += 1

    for txn in sorted(transactions, key=lambda x: x.get('date', '')):
        debit = txn['amount'] if txn.get('type') == 'DR' else 0
        credit = txn['amount'] if txn.get('type') == 'CR' else 0
        running_balance += debit - credit

        ws.cell(row=row, column=1, value=txn.get('date', '')).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=txn.get('description', '')).font = NORMAL_FONT
        if debit:
            ws.cell(row=row, column=3, value=debit).font = NORMAL_FONT
            ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        if credit:
            ws.cell(row=row, column=4, value=credit).font = NORMAL_FONT
            ws.cell(row=row, column=4).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=5, value=running_balance).font = NORMAL_FONT
        ws.cell(row=row, column=5).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=6, value=txn.get('account_code', '')).font = NORMAL_FONT
        acc_name = coa.get(txn.get('account_code', ''), {}).get('name', '')
        ws.cell(row=row, column=7, value=acc_name).font = NORMAL_FONT
        row += 1


def write_payroll_summary(wb, data):
    ws = wb.create_sheet("Payroll Summary")
    payroll = data.get('payroll', [])
    if not payroll:
        ws.cell(row=1, column=1, value="No payroll data provided").font = SUBTITLE_FONT
        return

    ws.cell(row=1, column=1, value="PAYROLL SUMMARY").font = TITLE_FONT
    headers = ["Month", "Employee", "Gross Salary", "EPF (ER)", "EPF (EE)",
               "SOCSO (ER)", "SOCSO (EE)", "EIS (ER)", "EIS (EE)", "PCB", "Net Pay"]
    row = write_header_row(ws, 3, headers, [12, 20, 14, 12, 12, 12, 12, 10, 10, 12, 14])

    totals = defaultdict(float)
    for entry in sorted(payroll, key=lambda x: (x.get('month', ''), x.get('employee', ''))):
        ws.cell(row=row, column=1, value=entry.get('month', '')).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=entry.get('employee', '')).font = NORMAL_FONT
        for col, key in enumerate(['gross', 'epf_er', 'epf_ee', 'socso_er', 'socso_ee',
                                    'eis_er', 'eis_ee', 'pcb', 'net_pay'], 3):
            val = entry.get(key, 0)
            ws.cell(row=row, column=col, value=val).font = NORMAL_FONT
            ws.cell(row=row, column=col).number_format = NUMBER_FORMAT
            totals[key] += val
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = HEADER_FONT
    for col, key in enumerate(['gross', 'epf_er', 'epf_ee', 'socso_er', 'socso_ee',
                                'eis_er', 'eis_ee', 'pcb', 'net_pay'], 3):
        cell = ws.cell(row=row, column=col, value=totals[key])
        cell.font = HEADER_FONT
        cell.number_format = NUMBER_FORMAT
        cell.border = DOUBLE_BORDER


def write_fixed_asset_register(wb, data):
    ws = wb.create_sheet("Fixed Asset Register")
    assets = data.get('fixed_assets', [])
    if not assets:
        ws.cell(row=1, column=1, value="No fixed assets").font = SUBTITLE_FONT
        return

    fy_end = data.get('company_info', {}).get('fy_end', '')

    ws.cell(row=1, column=1, value="FIXED ASSET REGISTER & DEPRECIATION SCHEDULE").font = TITLE_FONT
    headers = ["Description", "Category", "Date Acquired", "Cost (RM)", "Rate (%)",
               "Accum Depr B/F", "Current Year Depr", "Accum Depr C/F", "NBV"]
    row = write_header_row(ws, 3, headers, [30, 20, 14, 14, 10, 14, 14, 14, 14])

    total_cost = 0
    total_depr_bf = 0
    total_current_depr = 0
    total_depr_cf = 0
    total_nbv = 0

    for asset in sorted(assets, key=lambda x: x.get('category', '')):
        cost = asset.get('cost', 0)
        rate = asset.get('rate', 0)
        accum_bf = asset.get('accumulated_depr', 0)

        if fy_end and asset.get('date_acquired'):
            try:
                acq_date = datetime.strptime(asset['date_acquired'], '%Y-%m-%d')
                fy_end_date = datetime.strptime(fy_end, '%Y-%m-%d')
                fy_start_str = data.get('company_info', {}).get('fy_start', '')
                fy_start_date = datetime.strptime(fy_start_str, '%Y-%m-%d') if fy_start_str else None

                if acq_date > fy_end_date:
                    current_depr = 0
                elif fy_start_date and acq_date >= fy_start_date:
                    days_in_fy = (fy_end_date - fy_start_date).days + 1
                    days_used = (fy_end_date - acq_date).days + 1
                    current_depr = round(cost * (rate / 100) * (days_used / days_in_fy), 2)
                else:
                    remaining = cost - accum_bf
                    current_depr = min(round(cost * (rate / 100), 2), remaining)
            except (ValueError, TypeError):
                current_depr = round(cost * (rate / 100), 2)
        else:
            current_depr = round(cost * (rate / 100), 2)

        remaining = cost - accum_bf
        current_depr = min(current_depr, remaining)
        accum_cf = accum_bf + current_depr
        nbv = cost - accum_cf

        ws.cell(row=row, column=1, value=asset.get('description', '')).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=asset.get('category', '')).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=asset.get('date_acquired', '')).font = NORMAL_FONT
        for col, val in enumerate([cost, rate, accum_bf, current_depr, accum_cf, nbv], 4):
            ws.cell(row=row, column=col, value=val).font = NORMAL_FONT
            ws.cell(row=row, column=col).number_format = NUMBER_FORMAT
        row += 1

        total_cost += cost
        total_depr_bf += accum_bf
        total_current_depr += current_depr
        total_depr_cf += accum_cf
        total_nbv += nbv

    row += 1
    ws.cell(row=row, column=1, value="TOTAL").font = HEADER_FONT
    for col, val in enumerate([total_cost, None, total_depr_bf, total_current_depr, total_depr_cf, total_nbv], 4):
        if val is not None:
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = HEADER_FONT
            cell.number_format = NUMBER_FORMAT
            cell.border = DOUBLE_BORDER


def write_journal_entries(wb, data):
    ws = wb.create_sheet("Journal Entries")
    jes = data.get('journal_entries', [])

    ws.cell(row=1, column=1, value="JOURNAL ENTRIES").font = TITLE_FONT
    headers = ["JE #", "Date", "Account Code", "Account Name", "Description", "Debit (RM)", "Credit (RM)"]
    row = write_header_row(ws, 3, headers, [10, 12, 12, 30, 35, 14, 14])

    coa = data.get('chart_of_accounts', {})

    for je in sorted(jes, key=lambda x: x.get('je_number', '')):
        for line in je.get('lines', []):
            ws.cell(row=row, column=1, value=je.get('je_number', '')).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=je.get('date', '')).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=line.get('account_code', '')).font = NORMAL_FONT
            acc_name = coa.get(line.get('account_code', ''), {}).get('name', line.get('account_name', ''))
            ws.cell(row=row, column=4, value=acc_name).font = NORMAL_FONT
            ws.cell(row=row, column=5, value=je.get('description', '')).font = NORMAL_FONT
            if line.get('debit'):
                ws.cell(row=row, column=6, value=line['debit']).font = NORMAL_FONT
                ws.cell(row=row, column=6).number_format = NUMBER_FORMAT
            if line.get('credit'):
                ws.cell(row=row, column=7, value=line['credit']).font = NORMAL_FONT
                ws.cell(row=row, column=7).number_format = NUMBER_FORMAT
            row += 1
        row += 1


def build_general_ledger(data):
    """Build GL from journal entries, returning per-account transaction lists."""
    gl = defaultdict(list)
    coa = data.get('chart_of_accounts', {})
    opening = data.get('opening_balances', {})

    for code, amount in opening.items():
        if amount != 0:
            gl[code].append({
                'date': data.get('company_info', {}).get('fy_start', ''),
                'je_number': 'OB',
                'description': 'Opening Balance',
                'debit': amount if coa.get(code, {}).get('normal') == 'DR' else 0,
                'credit': amount if coa.get(code, {}).get('normal') == 'CR' else 0,
            })

    for je in data.get('journal_entries', []):
        for line in je.get('lines', []):
            code = line.get('account_code', '')
            gl[code].append({
                'date': je.get('date', ''),
                'je_number': je.get('je_number', ''),
                'description': je.get('description', ''),
                'debit': line.get('debit', 0),
                'credit': line.get('credit', 0),
            })

    return gl


def write_general_ledger(wb, data):
    ws = wb.create_sheet("General Ledger")
    gl = build_general_ledger(data)
    coa = data.get('chart_of_accounts', {})

    ws.cell(row=1, column=1, value="GENERAL LEDGER").font = TITLE_FONT
    row = 3

    for code in sorted(gl.keys()):
        if code not in coa:
            continue
        acc = coa[code]
        acc_name = acc.get('name', '')
        normal = acc.get('normal', 'DR')

        ws.cell(row=row, column=1, value=f"{code} - {acc_name}").font = HEADER_FONT
        row += 1
        headers = ["Date", "JE #", "Description", "Debit", "Credit", "Balance"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=h).font = HEADER_FONT
            ws.cell(row=row, column=col).fill = HEADER_FILL
        row += 1

        balance = 0
        for txn in sorted(gl[code], key=lambda x: x.get('date', '')):
            dr = txn.get('debit', 0) or 0
            cr = txn.get('credit', 0) or 0
            if normal == 'DR':
                balance += dr - cr
            else:
                balance += cr - dr

            ws.cell(row=row, column=1, value=txn.get('date', '')).font = NORMAL_FONT
            ws.cell(row=row, column=2, value=txn.get('je_number', '')).font = NORMAL_FONT
            ws.cell(row=row, column=3, value=txn.get('description', '')).font = NORMAL_FONT
            if dr:
                ws.cell(row=row, column=4, value=dr).font = NORMAL_FONT
                ws.cell(row=row, column=4).number_format = NUMBER_FORMAT
            if cr:
                ws.cell(row=row, column=5, value=cr).font = NORMAL_FONT
                ws.cell(row=row, column=5).number_format = NUMBER_FORMAT
            ws.cell(row=row, column=6, value=balance).font = NORMAL_FONT
            ws.cell(row=row, column=6).number_format = NUMBER_FORMAT
            row += 1

        ws.cell(row=row, column=3, value="Closing Balance").font = HEADER_FONT
        cell = ws.cell(row=row, column=6, value=balance)
        cell.font = HEADER_FONT
        cell.number_format = NUMBER_FORMAT
        cell.border = DOUBLE_BORDER
        row += 2

    set_column_widths(ws, [12, 10, 35, 14, 14, 14])


def compute_balances(data):
    """Compute closing balances for all accounts from GL."""
    gl = build_general_ledger(data)
    coa = data.get('chart_of_accounts', {})
    balances = {}

    for code in coa:
        normal = coa[code].get('normal', 'DR')
        balance = 0
        for txn in gl.get(code, []):
            dr = txn.get('debit', 0) or 0
            cr = txn.get('credit', 0) or 0
            if normal == 'DR':
                balance += dr - cr
            else:
                balance += cr - dr
        balances[code] = balance

    return balances


def write_trial_balance(wb, data):
    ws = wb.create_sheet("Trial Balance")
    coa = data.get('chart_of_accounts', {})
    balances = compute_balances(data)

    info = data.get('company_info', {})
    ws.cell(row=1, column=1, value=info.get('name', '')).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Trial Balance as at {info.get('fy_end', '')}").font = SUBTITLE_FONT

    headers = ["Code", "Account Name", "Debit (RM)", "Credit (RM)"]
    row = write_header_row(ws, 4, headers, [10, 40, 14, 14])

    total_dr = 0
    total_cr = 0

    for code in sorted(coa.keys()):
        bal = balances.get(code, 0)
        if bal == 0:
            continue
        normal = coa[code].get('normal', 'DR')

        ws.cell(row=row, column=1, value=code).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=coa[code].get('name', '')).font = NORMAL_FONT

        if (normal == 'DR' and bal > 0) or (normal == 'CR' and bal < 0):
            dr_val = abs(bal)
            ws.cell(row=row, column=3, value=dr_val).font = NORMAL_FONT
            ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
            total_dr += dr_val
        else:
            cr_val = abs(bal)
            ws.cell(row=row, column=4, value=cr_val).font = NORMAL_FONT
            ws.cell(row=row, column=4).number_format = NUMBER_FORMAT
            total_cr += cr_val
        row += 1

    row += 1
    ws.cell(row=row, column=2, value="TOTAL").font = HEADER_FONT
    cell_dr = ws.cell(row=row, column=3, value=total_dr)
    cell_dr.font = HEADER_FONT
    cell_dr.number_format = NUMBER_FORMAT
    cell_dr.border = DOUBLE_BORDER
    cell_cr = ws.cell(row=row, column=4, value=total_cr)
    cell_cr.font = HEADER_FONT
    cell_cr.number_format = NUMBER_FORMAT
    cell_cr.border = DOUBLE_BORDER

    row += 2
    diff = round(total_dr - total_cr, 2)
    ws.cell(row=row, column=2, value="Difference:").font = HEADER_FONT
    ws.cell(row=row, column=3, value=diff).font = HEADER_FONT
    if diff != 0:
        ws.cell(row=row, column=3).font = Font(name='Arial', size=10, bold=True, color='FF0000')


def write_income_statement(wb, data):
    ws = wb.create_sheet("Income Statement")
    coa = data.get('chart_of_accounts', {})
    balances = compute_balances(data)
    info = data.get('company_info', {})

    ws.cell(row=1, column=1, value=info.get('name', '')).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Statement of Comprehensive Income").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"For the financial year ended {info.get('fy_end', '')}").font = SUBTITLE_FONT
    set_column_widths(ws, [40, 14, 14])

    row = 5
    ws.cell(row=row, column=1, value="REVENUE").font = HEADER_FONT
    row += 1
    total_revenue = 0
    for code in sorted(coa.keys()):
        if code.startswith('4') and balances.get(code, 0) != 0:
            ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
            val = balances[code]
            ws.cell(row=row, column=2, value=val).font = NORMAL_FONT
            ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
            total_revenue += val
            row += 1
    ws.cell(row=row, column=1, value="Total Revenue").font = HEADER_FONT
    ws.cell(row=row, column=3, value=total_revenue).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=3).border = THIN_BORDER
    row += 2

    ws.cell(row=row, column=1, value="EXPENSES").font = HEADER_FONT
    row += 1
    total_expenses = 0
    for code in sorted(coa.keys()):
        if (code.startswith('5') or code.startswith('6')) and balances.get(code, 0) != 0:
            ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
            val = balances[code]
            ws.cell(row=row, column=2, value=val).font = NORMAL_FONT
            ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
            total_expenses += val
            row += 1
    ws.cell(row=row, column=1, value="Total Expenses").font = HEADER_FONT
    ws.cell(row=row, column=3, value=total_expenses).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=3).border = THIN_BORDER
    row += 2

    net_profit = total_revenue - total_expenses
    ws.cell(row=row, column=1, value="PROFIT / (LOSS) BEFORE TAX").font = HEADER_FONT
    ws.cell(row=row, column=3, value=net_profit).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 1

    tax = 0
    for code in sorted(coa.keys()):
        if code.startswith('9') and balances.get(code, 0) != 0:
            ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
            val = balances[code]
            ws.cell(row=row, column=2, value=val).font = NORMAL_FONT
            ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
            tax += val
            row += 1

    row += 1
    net_after_tax = net_profit - tax
    ws.cell(row=row, column=1, value="PROFIT / (LOSS) FOR THE YEAR").font = HEADER_FONT
    cell = ws.cell(row=row, column=3, value=net_after_tax)
    cell.font = HEADER_FONT
    cell.number_format = NUMBER_FORMAT
    cell.border = DOUBLE_BORDER


def write_balance_sheet(wb, data):
    ws = wb.create_sheet("Balance Sheet")
    coa = data.get('chart_of_accounts', {})
    balances = compute_balances(data)
    info = data.get('company_info', {})

    ws.cell(row=1, column=1, value=info.get('name', '')).font = TITLE_FONT
    ws.cell(row=2, column=1, value="Statement of Financial Position").font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"As at {info.get('fy_end', '')}").font = SUBTITLE_FONT
    set_column_widths(ws, [40, 14, 14])

    row = 5

    ws.cell(row=row, column=1, value="ASSETS").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Non-Current Assets").font = Font(name='Arial', size=10, bold=True, italic=True)
    row += 1
    total_nca = 0
    for code in sorted(coa.keys()):
        if code.startswith('13') or code.startswith('14') or code.startswith('16') or code.startswith('17'):
            bal = balances.get(code, 0)
            if bal != 0:
                ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
                ws.cell(row=row, column=2, value=bal).font = NORMAL_FONT
                ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
                total_nca += bal
                row += 1
    ws.cell(row=row, column=3, value=total_nca).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Current Assets").font = Font(name='Arial', size=10, bold=True, italic=True)
    row += 1
    total_ca = 0
    for code in sorted(coa.keys()):
        if code.startswith('10') or code.startswith('11') or code.startswith('12') or code.startswith('15'):
            bal = balances.get(code, 0)
            if bal != 0:
                ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
                ws.cell(row=row, column=2, value=bal).font = NORMAL_FONT
                ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
                total_ca += bal
                row += 1
    ws.cell(row=row, column=3, value=total_ca).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 1

    total_assets = total_nca + total_ca
    ws.cell(row=row, column=1, value="TOTAL ASSETS").font = HEADER_FONT
    cell = ws.cell(row=row, column=3, value=total_assets)
    cell.font = HEADER_FONT
    cell.number_format = NUMBER_FORMAT
    cell.border = DOUBLE_BORDER
    row += 2

    ws.cell(row=row, column=1, value="EQUITY AND LIABILITIES").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Equity").font = Font(name='Arial', size=10, bold=True, italic=True)
    row += 1
    total_equity = 0
    for code in sorted(coa.keys()):
        if code.startswith('3'):
            bal = balances.get(code, 0)
            if bal != 0:
                ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
                ws.cell(row=row, column=2, value=bal).font = NORMAL_FONT
                ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
                total_equity += bal
                row += 1
    ws.cell(row=row, column=3, value=total_equity).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Non-Current Liabilities").font = Font(name='Arial', size=10, bold=True, italic=True)
    row += 1
    total_ncl = 0
    for code in sorted(coa.keys()):
        if code.startswith('25') or code.startswith('26') or code.startswith('27'):
            bal = balances.get(code, 0)
            if bal != 0:
                ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
                ws.cell(row=row, column=2, value=bal).font = NORMAL_FONT
                ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
                total_ncl += bal
                row += 1
    ws.cell(row=row, column=3, value=total_ncl).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 1

    ws.cell(row=row, column=1, value="Current Liabilities").font = Font(name='Arial', size=10, bold=True, italic=True)
    row += 1
    total_cl = 0
    for code in sorted(coa.keys()):
        if code.startswith('20') or code.startswith('21') or code.startswith('22') or code.startswith('23') or code.startswith('24') or code.startswith('28') or code.startswith('29'):
            bal = balances.get(code, 0)
            if bal != 0:
                ws.cell(row=row, column=1, value=f"  {coa[code]['name']}").font = NORMAL_FONT
                ws.cell(row=row, column=2, value=bal).font = NORMAL_FONT
                ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
                total_cl += bal
                row += 1
    ws.cell(row=row, column=3, value=total_cl).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 1

    total_el = total_equity + total_ncl + total_cl
    ws.cell(row=row, column=1, value="TOTAL EQUITY AND LIABILITIES").font = HEADER_FONT
    cell = ws.cell(row=row, column=3, value=total_el)
    cell.font = HEADER_FONT
    cell.number_format = NUMBER_FORMAT
    cell.border = DOUBLE_BORDER

    row += 2
    diff = round(total_assets - total_el, 2)
    if diff != 0:
        ws.cell(row=row, column=1, value="*** IMBALANCE ***").font = Font(name='Arial', size=10, bold=True, color='FF0000')
        ws.cell(row=row, column=3, value=diff).font = Font(name='Arial', size=10, bold=True, color='FF0000')


def write_tax_computation(wb, data):
    ws = wb.create_sheet("Tax Computation")
    tax = data.get('tax_computation', {})
    info = data.get('company_info', {})

    if not tax:
        ws.cell(row=1, column=1, value="Tax computation data not provided").font = SUBTITLE_FONT
        return

    ws.cell(row=1, column=1, value=info.get('name', '')).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Tax Computation for YA {info.get('fy_end', '')[:4] if info.get('fy_end') else ''}").font = SUBTITLE_FONT
    set_column_widths(ws, [40, 14, 14])

    row = 4
    ws.cell(row=row, column=1, value="Net Profit per Accounts").font = NORMAL_FONT
    ws.cell(row=row, column=3, value=tax.get('net_profit', 0)).font = NORMAL_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 2

    ws.cell(row=row, column=1, value="Add: Non-Deductible Expenses").font = HEADER_FONT
    row += 1
    total_add = 0
    for item in tax.get('add_backs', []):
        ws.cell(row=row, column=1, value=f"  {item.get('description', '')}").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=item.get('amount', 0)).font = NORMAL_FONT
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        total_add += item.get('amount', 0)
        row += 1
    ws.cell(row=row, column=3, value=total_add).font = NORMAL_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 2

    ws.cell(row=row, column=1, value="Less: Non-Taxable Income").font = HEADER_FONT
    row += 1
    total_less = 0
    for item in tax.get('less_items', []):
        ws.cell(row=row, column=1, value=f"  {item.get('description', '')}").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=item.get('amount', 0)).font = NORMAL_FONT
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        total_less += item.get('amount', 0)
        row += 1
    ws.cell(row=row, column=3, value=total_less).font = NORMAL_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    row += 2

    adjusted_income = tax.get('net_profit', 0) + total_add - total_less
    ws.cell(row=row, column=1, value="Adjusted Income").font = HEADER_FONT
    ws.cell(row=row, column=3, value=adjusted_income).font = HEADER_FONT
    ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
    ws.cell(row=row, column=3).border = THIN_BORDER
    row += 2

    ca = tax.get('capital_allowances', {})
    if ca:
        ws.cell(row=row, column=1, value="Less: Capital Allowances").font = HEADER_FONT
        row += 1
        ws.cell(row=row, column=1, value="  Balance b/f").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=ca.get('bf', 0)).font = NORMAL_FONT
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        row += 1
        ws.cell(row=row, column=1, value="  Initial Allowance").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=ca.get('ia', 0)).font = NORMAL_FONT
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        row += 1
        ws.cell(row=row, column=1, value="  Annual Allowance").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=ca.get('aa', 0)).font = NORMAL_FONT
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        row += 1
        total_ca_available = ca.get('bf', 0) + ca.get('ia', 0) + ca.get('aa', 0)
        absorbed = min(ca.get('absorbed', total_ca_available), adjusted_income, total_ca_available)
        ws.cell(row=row, column=1, value="  Absorbed").font = NORMAL_FONT
        ws.cell(row=row, column=3, value=absorbed).font = NORMAL_FONT
        ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        row += 1
        cf = total_ca_available - absorbed
        ws.cell(row=row, column=1, value="  Balance c/f").font = NORMAL_FONT
        ws.cell(row=row, column=2, value=cf).font = NORMAL_FONT
        ws.cell(row=row, column=2).number_format = NUMBER_FORMAT
        row += 2

        chargeable = max(adjusted_income - absorbed, 0)
    else:
        chargeable = adjusted_income

    ws.cell(row=row, column=1, value="Chargeable Income").font = HEADER_FONT
    cell = ws.cell(row=row, column=3, value=chargeable)
    cell.font = HEADER_FONT
    cell.number_format = NUMBER_FORMAT
    cell.border = DOUBLE_BORDER


def write_queries(wb, data):
    ws = wb.create_sheet("Queries & Notes")
    queries = data.get('queries', [])

    ws.cell(row=1, column=1, value="QUERIES & OUTSTANDING ITEMS").font = TITLE_FONT
    if not queries:
        ws.cell(row=3, column=1, value="No outstanding queries. All items resolved.").font = SUBTITLE_FONT
        return

    headers = ["#", "Date", "Amount (RM)", "Description", "Question", "Status"]
    row = write_header_row(ws, 3, headers, [5, 12, 14, 30, 40, 10])

    for i, q in enumerate(queries, 1):
        ws.cell(row=row, column=1, value=i).font = NORMAL_FONT
        ws.cell(row=row, column=2, value=q.get('date', '')).font = NORMAL_FONT
        ws.cell(row=row, column=3, value=q.get('amount', 0)).font = NORMAL_FONT
        ws.cell(row=row, column=3).number_format = NUMBER_FORMAT
        ws.cell(row=row, column=4, value=q.get('description', '')).font = NORMAL_FONT
        ws.cell(row=row, column=5, value=q.get('question', '')).font = NORMAL_FONT
        ws.cell(row=row, column=6, value=q.get('status', 'open')).font = NORMAL_FONT
        row += 1


def generate_workbook(data, output_path):
    """Generate complete working papers workbook from classified data."""
    wb = Workbook()
    wb.remove(wb.active)

    write_company_info(wb, data)
    write_chart_of_accounts(wb, data)
    write_bank_transactions(wb, data)
    write_payroll_summary(wb, data)
    write_fixed_asset_register(wb, data)
    write_journal_entries(wb, data)
    write_general_ledger(wb, data)
    write_trial_balance(wb, data)
    write_income_statement(wb, data)
    write_balance_sheet(wb, data)
    write_tax_computation(wb, data)
    write_queries(wb, data)

    wb.save(output_path)
    print(f"Workbook generated: {output_path}")

    balances = compute_balances(data)
    coa = data.get('chart_of_accounts', {})
    total_dr = sum(abs(b) for c, b in balances.items() if b != 0 and ((coa.get(c, {}).get('normal') == 'DR' and b > 0) or (coa.get(c, {}).get('normal') == 'CR' and b < 0)))
    total_cr = sum(abs(b) for c, b in balances.items() if b != 0 and ((coa.get(c, {}).get('normal') == 'CR' and b > 0) or (coa.get(c, {}).get('normal') == 'DR' and b < 0)))
    diff = round(total_dr - total_cr, 2)
    if diff == 0:
        print("TB CHECK: BALANCED")
    else:
        print(f"TB CHECK: IMBALANCED by RM{diff}")

    return output_path


def main():
    parser = argparse.ArgumentParser(description='Generate accounting working papers from classified data')
    parser.add_argument('input_json', help='Path to classified data JSON file')
    parser.add_argument('output_xlsx', help='Path for output Excel workbook')
    parser.add_argument('--entity-type', choices=['sdn_bhd', 'bhd', 'plt', 'sole_prop', 'partnership', 'koperasi'],
                       default='sdn_bhd', help='Entity type (default: sdn_bhd)')
    args = parser.parse_args()

    with open(args.input_json, 'r') as f:
        data = json.load(f)

    generate_workbook(data, args.output_xlsx)


if __name__ == '__main__':
    main()
