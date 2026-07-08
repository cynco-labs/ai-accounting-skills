#!/usr/bin/env python3
# SPDX-License-Identifier: Apache-2.0
"""Extract Maybank Islamic (and compatible Maybank) e-statement PDFs to Excel/CSV/JSON.

Proven method (CJT Bakery-style statements):
  pdfplumber text → line regex → Decimal running-balance proof → openpyxl/CSV

Usage:
  python3 scripts/extract_maybank_islamic_pdf.py \\
    --input /path/to/folder-or.pdf \\
    --output /path/to/out.xlsx

  python3 scripts/extract_maybank_islamic_pdf.py \\
    --input ./statements --output ./txns.xlsx --also-json ./txns.json

Requires: pdfplumber, openpyxl
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from decimal import Decimal
from pathlib import Path

try:
    import pdfplumber
except ImportError:
    print("ERROR: pip install pdfplumber", file=sys.stderr)
    sys.exit(2)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl", file=sys.stderr)
    sys.exit(2)


LINE_RE = re.compile(
    r"^(\d{2}/\d{2})\s+(.+?)\s+([\d,]+\.\d{2})([+-])\s+([\d,]+\.\d{2})(DR)?\s*$"
)
BEGIN_RE = re.compile(r"BEGINNING BALANCE\s+([\d,]+\.\d{2})(DR)?")
STMT_DATE_RE = re.compile(r"結單日期\s*:\s*(\d{2}/\d{2}/\d{2})")
ACCOUNT_RE = re.compile(r"戶號\s*:\s*(\d+)")
ENTITY_RE = re.compile(r"^(CJT BAKERY SDN\. BHD\.|.+SDN\.?\s*BHD\.?)", re.M)

STOP_DETAIL = re.compile(
    r"ENDING BALANCE|TOTAL DEBIT|TOTAL CREDIT|PROFIT OUTSTANDING|BAKI LEGAR|"
    r"LEDGER\s*=|Perhatian|NOT PROTECTED|Maybank Islamic|PAGE\s*:|MUKA/|"
    r"All items and balances|Please notify|EXCHANGE YOUR CURRENCY|"
    r"VISIT WWW|PLEASE BE REMINDED|CREDIT TO MULTIPLE|EFFECTIVE 1|"
    r"notified in writing|THANK YOU FOR YOUR|NOTICE:|KINDLY BE INFORMED|"
    r"STARTING 4TH|BERMULA|KINI, ANDA|INWARD RETURN|NOTIS PEMULANGAN|"
    r"HOUSE CHEQUES|STAMP DUTY|ECONFIRMATION|ANNOUCEMENTS",
    re.I,
)

SKIP_LINE = re.compile(
    r"^(Maybank|15th Floor|IBS AEON|MUKA/|TARIKH PENYATA|NO 31-M|"
    r"JALAN TENGKU|NOMBOR AKAUN|SEKSYEN 9|40100 SHAH|ACCOUNT|NOT PROTECTED|"
    r"URUSNIAGA|TARIKH MASUK|ENTRY DATE|BAKI LEGAR|LEDGER|BALANCE|Perhatian|"
    r"\(1\)|\(2\)|Sila beritahu|All items|Please notify|Overdrawn|Semua maklumat|"
    r"STATEMENT DATE|VALUE DATE|TRANSACTION|ENDING BALANCE|TOTAL DEBIT|"
    r"TOTAL CREDIT|PROFIT OUTSTANDING|FCN|EXCHANGE YOUR|VISIT |PLEASE |"
    r"CREDIT TO|EFFECTIVE |KINI,|NOTICE:|KINDLY |STARTING |BERMULA|"
    r"INWARD |NOTIS |HOUSE |STAMP |THANK YOU|ADVISED TO|NOTE:|FOR ANY|"
    r"WE WOULD|PUBLIC ISSUE|NEAREST BRANCH|OR CALL|HTTP://|WWW\.|"
    r"1-300|1300|RM50,000|RM250,000|NOT A M2U|IF YOU ARE|THE CHECKER|"
    r"SINGLE ACCOUNT|DAILY FINANCIAL|M2U|MAE APP|MAYBANK2U|MONEY EXCHANGE|"
    r"HIDDEN|CHARGES!|CEK |CHEQUE|GAZETTED|FINANCE ACT|REVISED FROM|"
    r"PER CHEQUE|FURTHER CLARIFICATIONS|ANNOUCEMENTS|UNDERSTANDING|"
    r"COURIERED|BRANCHES\.|MAILING ADDRESS|SERVICE DISRUPTIONS|"
    r"DIDEPOSITKAN|DIKREDITKAN|HARI BEKERJA|PEMINDAHAN DANA|"
    r"DISCONTINUING|STANDING INSTRUCTION|PHYSICAL LETTERS|"
    r"ENGAGE YOUR AUDITOR|RESPONSIBLE OR LIABLE|RETURN CHEQUE|"
    r"SME FIRST|ENTRY DATE VALUE DATE|TRANSACTION DESCRIPTION)",
    re.I,
)


def money(s: str) -> Decimal:
    return Decimal(s.replace(",", "")).quantize(Decimal("0.01"))


def year_from_stmt(stmt_date: str) -> tuple[int, int, int]:
    d, m, y = stmt_date.split("/")
    return 2000 + int(y), int(m), int(d)


def full_date(entry_ddmm: str, stmt_year: int) -> str:
    dd, mm = entry_ddmm.split("/")
    return f"{stmt_year}-{mm}-{dd}"


def clean_detail(parts: list[str]) -> str:
    out: list[str] = []
    seen: set[str] = set()
    for p in parts:
        p = p.strip()
        if not p or p in ("*", "__STOP__"):
            continue
        if STOP_DETAIL.search(p) or "notified in writing" in p.lower():
            break
        if SKIP_LINE.match(p) or len(p) > 120:
            continue
        if p not in seen:
            seen.add(p)
            out.append(p)
        if len(out) >= 6:
            break
    return " | ".join(out)


def extract_pdf(pdf_path: Path) -> tuple[dict, list[dict]]:
    meta: dict = {
        "file": pdf_path.name,
        "pages": 0,
        "statement_date": None,
        "beginning_balance": None,
        "account_number": None,
        "entity": None,
        "errors": [],
        "bank_footer_debit": None,
        "bank_footer_credit": None,
        "bank_footer_ending": None,
    }
    pages: list[tuple[int, str]] = []
    with pdfplumber.open(pdf_path) as doc:
        meta["pages"] = len(doc.pages)
        for pi, page in enumerate(doc.pages, 1):
            text = page.extract_text() or ""
            pages.append((pi, text))
            if pi == 1:
                m = STMT_DATE_RE.search(text)
                if m:
                    meta["statement_date"] = m.group(1)
                m2 = ACCOUNT_RE.search(text)
                if m2:
                    meta["account_number"] = m2.group(1)
                m3 = ENTITY_RE.search(text)
                if m3:
                    meta["entity"] = m3.group(1).strip()
                mb = BEGIN_RE.search(text)
                if mb:
                    meta["beginning_balance"] = money(mb.group(1))
                    if mb.group(2):
                        meta["beginning_balance"] = -meta["beginning_balance"]
            md = re.search(r"TOTAL DEBIT\s*:\s*([\d,]+\.\d{2})", text)
            mc = re.search(r"TOTAL CREDIT\s*:\s*([\d,]+\.\d{2})", text)
            me = re.search(r"ENDING BALANCE\s*:\s*([\d,]+\.\d{2})", text)
            if md:
                meta["bank_footer_debit"] = money(md.group(1))
            if mc:
                meta["bank_footer_credit"] = money(mc.group(1))
            if me:
                meta["bank_footer_ending"] = money(me.group(1))

    if not meta["statement_date"] or meta["beginning_balance"] is None:
        meta["errors"].append("Missing statement date or beginning balance — not a supported Maybank layout?")
        return meta, []

    stmt_year, stmt_month, stmt_day = year_from_stmt(meta["statement_date"])
    meta["period"] = f"{stmt_year}-{stmt_month:02d}"
    meta["statement_date_iso"] = f"{stmt_year}-{stmt_month:02d}-{stmt_day:02d}"

    rows: list[dict] = []
    current: dict | None = None
    pending: list[str] = []

    def flush() -> None:
        nonlocal current, pending
        if current:
            current["description_detail"] = clean_detail(pending)
            rows.append(current)
            current = None
            pending = []

    for pi, text in pages:
        for raw in text.splitlines():
            line = raw.strip()
            if not line:
                continue
            if re.match(r"^ENDING BALANCE\s*:", line, re.I):
                flush()
                continue
            if re.match(r"^TOTAL (DEBIT|CREDIT)\s*:", line, re.I):
                continue
            if SKIP_LINE.match(line):
                continue
            if STOP_DETAIL.search(line) and not LINE_RE.match(line):
                if current and pending:
                    pending.append("__STOP__")
                continue

            m = LINE_RE.match(line)
            if m:
                flush()
                ddmm, typ, amt, sign, bal, dr = m.groups()
                amount = money(amt)
                balance = money(bal)
                if dr:
                    balance = -balance
                signed = amount if sign == "+" else -amount
                current = {
                    "date": full_date(ddmm, stmt_year),
                    "entry_date": ddmm,
                    "type": typ.strip(),
                    "amount": amount,
                    "signed_amount": signed,
                    "direction": "inflow" if sign == "+" else "outflow",
                    "balance": balance,
                    "description_detail": "",
                    "source_file": pdf_path.name,
                    "source_page": pi,
                    "period": meta["period"],
                    "account_number": meta.get("account_number"),
                }
                pending = []
                continue

            if current is not None:
                if pending and pending[-1] == "__STOP__":
                    continue
                if STOP_DETAIL.search(line) or SKIP_LINE.match(line) or len(line) > 100:
                    if STOP_DETAIL.search(line):
                        pending.append("__STOP__")
                    continue
                pending.append(line)
    flush()

    prev = meta["beginning_balance"]
    for i, r in enumerate(rows):
        expected = (prev + r["signed_amount"]).quantize(Decimal("0.01"))
        if expected != r["balance"]:
            meta["errors"].append(
                f"Balance break #{i+1} {r['date']} {r['type']}: expected {expected} got {r['balance']}"
            )
            if len(meta["errors"]) > 20:
                break
        prev = r["balance"]

    meta["ending_balance"] = rows[-1]["balance"] if rows else None
    meta["txn_count"] = len(rows)
    meta["total_inflow"] = sum((r["amount"] for r in rows if r["direction"] == "inflow"), Decimal("0"))
    meta["total_outflow"] = sum((r["amount"] for r in rows if r["direction"] == "outflow"), Decimal("0"))
    if meta["beginning_balance"] is not None and meta["ending_balance"] is not None:
        net = meta["total_inflow"] - meta["total_outflow"]
        recon = (meta["beginning_balance"] + net - meta["ending_balance"]).quantize(Decimal("0.01"))
        meta["open_close_diff"] = recon
        meta["line_balance_ok"] = len(meta["errors"]) == 0
        meta["open_close_ok"] = recon == 0
    return meta, rows


def collect_inputs(path: Path) -> list[Path]:
    if path.is_file() and path.suffix.lower() == ".pdf":
        return [path]
    if path.is_dir():
        return sorted(path.glob("*.pdf")) + sorted(path.glob("*.PDF"))
    return []


def write_excel(metas: list[dict], all_rows: list[dict], out: Path) -> None:
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    title_font = Font(name="Arial", bold=True, size=14)
    normal = Font(name="Arial", size=10)
    blue = Font(name="Arial", size=10, color="0000FF")
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    money_fmt = '#,##0.00;(#,##0.00);"-"'
    in_fill = PatternFill("solid", fgColor="E2EFDA")
    out_fill = PatternFill("solid", fgColor="FCE4D6")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    tx_headers = [
        "Date", "Entry date (DD/MM)", "Period", "Type", "Description / counterparty detail",
        "Direction", "Amount (MYR)", "Signed amount (MYR)", "Statement balance (MYR)",
        "Source file", "Source page", "Account number",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    entity = next((m.get("entity") for m in metas if m.get("entity")), "Unknown entity")
    acct = next((m.get("account_number") for m in metas if m.get("account_number")), "")
    ws["A1"] = f"{entity} — Maybank Islamic statement extract"
    ws["A1"].font = title_font
    ws["A2"] = f"Account {acct} | Method: pdfplumber + line regex + Decimal balance proof"
    ws["A2"].font = normal
    ws["A3"] = "Do not invent rows. FAIL on balance break. Prefer bank CSV when available."
    ws["A3"].font = normal

    headers = [
        "File", "Statement date", "Period", "Pages", "Begin", "End", "Txns",
        "Inflow", "Outflow", "Net", "Open→close", "Line balances",
        "Footer CREDIT", "Footer DEBIT",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(5, c, h)
        cell.font = header_font
        cell.fill = header_fill

    for i, meta in enumerate(metas):
        r = 6 + i
        begin, end = meta.get("beginning_balance"), meta.get("ending_balance")
        tin, tout = meta.get("total_inflow") or 0, meta.get("total_outflow") or 0
        net = (tin - tout) if begin is not None else None
        vals = [
            meta["file"], meta.get("statement_date_iso"), meta.get("period"), meta.get("pages"),
            float(begin) if begin is not None else None,
            float(end) if end is not None else None,
            meta.get("txn_count", 0),
            float(tin) if tin is not None else None,
            float(tout) if tout is not None else None,
            float(net) if net is not None else None,
            "PASS" if meta.get("open_close_ok") else f"FAIL {meta.get('open_close_diff')}",
            "PASS" if meta.get("line_balance_ok") else f"FAIL ({len(meta.get('errors') or [])})",
            float(meta["bank_footer_credit"]) if meta.get("bank_footer_credit") is not None else None,
            float(meta["bank_footer_debit"]) if meta.get("bank_footer_debit") is not None else None,
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = blue if c in (5, 6, 8, 9, 10, 13, 14) else normal
            cell.border = thin
            if c in (5, 6, 8, 9, 10, 13, 14) and isinstance(v, float):
                cell.number_format = money_fmt
            if c in (11, 12) and isinstance(v, str) and not str(v).startswith("PASS"):
                cell.fill = warn_fill

    for i, w in enumerate([22, 12, 10, 8, 14, 14, 8, 12, 12, 12, 12, 12, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    def write_txns(sheet, rows: list[dict], title: str, sub: str) -> None:
        sheet["A1"] = title
        sheet["A1"].font = title_font
        sheet["A2"] = sub
        sheet["A2"].font = normal
        start = 4
        for c, h in enumerate(tx_headers, 1):
            cell = sheet.cell(start, c, h)
            cell.font = header_font
            cell.fill = header_fill
        for i, row in enumerate(rows, start + 1):
            vals = [
                row["date"], row["entry_date"], row["period"], row["type"],
                row.get("description_detail") or "",
                row["direction"], float(row["amount"]), float(row["signed_amount"]),
                float(row["balance"]), row["source_file"], row["source_page"],
                row.get("account_number") or "",
            ]
            for c, v in enumerate(vals, 1):
                cell = sheet.cell(i, c, v)
                cell.font = blue if c in (7, 8, 9) else normal
                cell.border = thin
                if c in (7, 8, 9):
                    cell.number_format = money_fmt
                if c == 6:
                    cell.fill = in_fill if v == "inflow" else out_fill
        end = start + len(rows)
        sheet.cell(end + 2, 6, "Sum inflows").font = Font(name="Arial", bold=True)
        sheet.cell(end + 2, 7, f'=SUMIF(F{start+1}:F{end},"inflow",G{start+1}:G{end})')
        sheet.cell(end + 2, 7).number_format = money_fmt
        sheet.cell(end + 3, 6, "Sum outflows").font = Font(name="Arial", bold=True)
        sheet.cell(end + 3, 7, f'=SUMIF(F{start+1}:F{end},"outflow",G{start+1}:G{end})')
        sheet.cell(end + 3, 7).number_format = money_fmt
        for i, w in enumerate([12, 12, 10, 28, 48, 10, 12, 14, 16, 16, 8, 14], 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        sheet.freeze_panes = "A5"
        if rows:
            sheet.auto_filter.ref = f"A{start}:L{end}"

    ws_all = wb.create_sheet("All Transactions")
    write_txns(ws_all, all_rows, "All transactions", f"{len(all_rows)} lines | sorted by period/date")

    by_period: dict[str, list] = defaultdict(list)
    for r in all_rows:
        by_period[r["period"]].append(r)
    for period in sorted(by_period.keys()):
        meta = next((m for m in metas if m.get("period") == period), {})
        sh = wb.create_sheet(period[:31])
        write_txns(
            sh,
            by_period[period],
            f"{period} — {meta.get('file', '')}",
            f"Begin {meta.get('beginning_balance')} → End {meta.get('ending_balance')} | "
            f"line={'PASS' if meta.get('line_balance_ok') else 'FAIL'}",
        )

    wsq = wb.create_sheet("QA_Checks")
    wsq["A1"] = "QA"
    wsq["A1"].font = title_font
    wsq["A3"] = "Total transactions"
    wsq["B3"] = len(all_rows)
    wsq["A4"] = "Files"
    wsq["B4"] = len(metas)
    wsq["A6"] = "Per-file status"
    wsq["A6"].font = Font(name="Arial", bold=True)
    wsq["A7"] = "File"
    wsq["B7"] = "Line balance"
    wsq["C7"] = "Open→close"
    wsq["D7"] = "Errors"
    for i, m in enumerate(metas, 8):
        wsq.cell(i, 1, m["file"])
        wsq.cell(i, 2, "PASS" if m.get("line_balance_ok") else "FAIL")
        wsq.cell(i, 3, "PASS" if m.get("open_close_ok") else "FAIL")
        wsq.cell(i, 4, "; ".join((m.get("errors") or [])[:3]) or "OK")
    wsq.column_dimensions["A"].width = 24
    wsq.column_dimensions["D"].width = 80

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def rows_to_schema(all_rows: list[dict], metas: list[dict], client_slug: str) -> dict:
    """transactions.schema.json-compatible payload."""
    banks = []
    seen = set()
    for m in metas:
        aid = m.get("account_number") or "unknown"
        if aid in seen:
            continue
        seen.add(aid)
        banks.append({
            "id": aid,
            "name": "Maybank Islamic",
            "account_number_last4": str(aid)[-4:] if aid else None,
        })
    txns = []
    for i, r in enumerate(all_rows, 1):
        txns.append({
            "id": f"txn-{r['date'].replace('-', '')}-{i:05d}",
            "date": r["date"],
            "description": f"{r['type']}" + (f" | {r['description_detail']}" if r.get("description_detail") else ""),
            "amount": float(r["amount"]),
            "direction": r["direction"],
            "bank_account_id": r.get("account_number") or "unknown",
            "running_balance": float(r["balance"]),
            "account_code": None,
            "classification_basis": None,
            "source_file": r["source_file"],
            "source_ref": f"page:{r['source_page']}",
        })
    return {
        "schema_version": "0.0.1",
        "client_slug": client_slug,
        "currency": "MYR",
        "bank_accounts": banks,
        "transactions": txns,
        "extraction_meta": [
            {
                "file": m["file"],
                "period": m.get("period"),
                "txn_count": m.get("txn_count"),
                "line_balance_ok": m.get("line_balance_ok"),
                "open_close_ok": m.get("open_close_ok"),
                "errors": m.get("errors") or [],
            }
            for m in metas
        ],
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="Extract Maybank Islamic e-statement PDFs")
    ap.add_argument("--input", type=Path, required=True, help="PDF file or folder of PDFs")
    ap.add_argument("--output", type=Path, required=True, help="Output .xlsx path")
    ap.add_argument("--also-json", type=Path, default=None, help="Optional transactions.json")
    ap.add_argument("--client-slug", default="client")
    ap.add_argument("--fail-on-error", action="store_true", help="Exit 1 if any balance check fails")
    args = ap.parse_args()

    files = collect_inputs(args.input)
    if not files:
        print(f"ERROR: no PDFs under {args.input}", file=sys.stderr)
        return 1

    metas: list[dict] = []
    all_rows: list[dict] = []
    for f in files:
        print(f"Parsing {f.name} ...", flush=True)
        meta, rows = extract_pdf(f)
        metas.append(meta)
        all_rows.extend(rows)
        status = "PASS" if meta.get("line_balance_ok") and meta.get("open_close_ok") else "CHECK"
        print(
            f"  {status} period={meta.get('period')} txns={meta.get('txn_count')} "
            f"begin={meta.get('beginning_balance')} end={meta.get('ending_balance')} "
            f"errors={len(meta.get('errors') or [])}",
            flush=True,
        )

    all_rows.sort(key=lambda r: (r["period"], r["date"], r["source_page"], r["entry_date"]))
    metas.sort(key=lambda m: m.get("period") or "")

    write_excel(metas, all_rows, args.output)
    print(f"Wrote {args.output} ({len(all_rows)} transactions)")

    if args.also_json:
        payload = rows_to_schema(all_rows, metas, args.client_slug)
        args.also_json.parent.mkdir(parents=True, exist_ok=True)
        args.also_json.write_text(json.dumps(payload, indent=2) + "\n", encoding="utf-8")
        print(f"Wrote {args.also_json}")

    failed = any(not (m.get("line_balance_ok") and m.get("open_close_ok")) for m in metas)
    if args.fail_on_error and failed:
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
